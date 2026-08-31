from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "arcus-aws-full-system-costs.xlsx"

GPT_MODEL = "GPT-5 mini"
GPT_INPUT_PER_M = 0.25
GPT_OUTPUT_PER_M = 2.00
AWS_REGION = "AWS US East (N. Virginia)"
AWS_PRICING_BASIS = "On-demand, Linux — one server running the full system (portal, API, database, files)"


def cost_per_case(input_tokens: int, output_tokens: int) -> float:
    return (input_tokens / 1_000_000) * GPT_INPUT_PER_M + (output_tokens / 1_000_000) * GPT_OUTPUT_PER_M


# Practical AWS: single EC2 + EBS, same layout as a VPS (Docker, everything on one box).
EBS_GP3_GB = 0.08


def aws_hosting_lines(compute_monthly: float, storage_gb: int) -> list[dict[str, str | float]]:
    storage = round(storage_gb * EBS_GP3_GB, 2)
    lines = [
        {
            "Cost Item": "Server",
            "What It Covers": "Runs the full Arcus system — portal, admin, API, and database",
            "Monthly Cost (USD)": compute_monthly,
        },
        {
            "Cost Item": "Storage",
            "What It Covers": "Database, uploaded documents, and report files on the same server",
            "Monthly Cost (USD)": storage,
        },
    ]
    total = round(compute_monthly + storage, 2)
    lines.append(
        {
            "Cost Item": "AWS hosting total",
            "What It Covers": "Estimated monthly cost to run the full system on AWS",
            "Monthly Cost (USD)": total,
        }
    )
    return lines


CLIENT_SCENARIOS = [
    {
        "Client Setup": "200 LP portal clients - all have portal access",
        "Why This Size": "Higher portal traffic — regular logins, downloads, notices, reports, and quarter-end spikes.",
        # m5.xlarge on-demand us-east-1: 4 vCPU / 16 GB — enough for full stack on one server
        "aws_lines": aws_hosting_lines(140.16, storage_gb=250),
    },
    {
        "Client Setup": "100 LP portal clients + 100 clients who only receive reports by email",
        "Why This Size": "Lower portal traffic — half the LP base logs in; the rest mainly receive emailed reports.",
        # m5.large on-demand us-east-1: 2 vCPU / 8 GB
        "aws_lines": aws_hosting_lines(70.08, storage_gb=200),
    },
]


AI_LINES = [
    {
        "AI Use": "KYC onboarding",
        "What it covers": "Reading application documents, summarising risk points, and preparing compliance-ready notes",
        "Estimated Cost": round(cost_per_case(120_000, 8_000), 2),
    },
    {
        "AI Use": "Portfolio applicant due diligence",
        "What it covers": "Reviewing DDQ packs, evidence, gaps, and producing a due-diligence summary",
        "Estimated Cost": round(cost_per_case(300_000, 20_000), 2),
    },
    {
        "AI Use": "Reading investee financials",
        "What it covers": "Reading submitted financial packs and generating KPI / variance commentary",
        "Estimated Cost": round(cost_per_case(180_000, 12_000), 2),
    },
    {
        "AI Use": "Writing reports / reviews / analyst notes",
        "What it covers": "Drafting polished narrative output for reviews, reports, and analyst commentary",
        "Estimated Cost": round(cost_per_case(60_000, 10_000), 2),
    },
]


MONTHLY_AI_BUNDLES = [
    {
        "Usage Month": "Light month",
        "Assumed Activity": "10 KYC cases, 6 DD cases, 20 investee financial packs, 40 report/review drafts",
    },
    {
        "Usage Month": "Active month",
        "Assumed Activity": "20 KYC cases, 12 DD cases, 40 investee financial packs, 100 report/review drafts",
    },
    {
        "Usage Month": "Busy quarter-end month",
        "Assumed Activity": "40 KYC cases, 25 DD cases, 80 investee financial packs, 200 report/review drafts",
    },
]

line_lookup = {row["AI Use"]: row["Estimated Cost"] for row in AI_LINES}
for row in MONTHLY_AI_BUNDLES:
    if row["Usage Month"] == "Light month":
        total = (
            10 * line_lookup["KYC onboarding"]
            + 6 * line_lookup["Portfolio applicant due diligence"]
            + 20 * line_lookup["Reading investee financials"]
            + 40 * line_lookup["Writing reports / reviews / analyst notes"]
        )
    elif row["Usage Month"] == "Active month":
        total = (
            20 * line_lookup["KYC onboarding"]
            + 12 * line_lookup["Portfolio applicant due diligence"]
            + 40 * line_lookup["Reading investee financials"]
            + 100 * line_lookup["Writing reports / reviews / analyst notes"]
        )
    else:
        total = (
            40 * line_lookup["KYC onboarding"]
            + 25 * line_lookup["Portfolio applicant due diligence"]
            + 80 * line_lookup["Reading investee financials"]
            + 200 * line_lookup["Writing reports / reviews / analyst notes"]
        )
    row["Estimated GPT Cost / Month (USD)"] = round(total, 2)


def aws_total(scenario: dict) -> float:
    for row in scenario["aws_lines"]:
        if row["Cost Item"] == "AWS hosting total":
            return float(row["Monthly Cost (USD)"])
    return 0.0


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="E8F0FE")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row_num: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_table(ws, start_row: int, end_row: int, cols: int, total_rows: set[int] | None = None):
    total_rows = total_rows or set()
    for row_idx in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if row_idx in total_rows:
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)


def fit_columns(ws, widths: dict[int, int]):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_section_title(ws, row: int, title: str) -> int:
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = Font(size=13, bold=True)
    return row + 1


wb = Workbook()

ws = wb.active
ws.title = "Full System Costs"
ws["A1"] = "Arcus — Full System Monthly Cost Estimate (AWS)"
ws["A1"].font = Font(size=16, bold=True)
ws["A2"] = "Rough monthly costs to run the full Arcus system on AWS — one server per client setup, same approach as a standard VPS deployment."
ws["A2"].alignment = Alignment(wrap_text=True)

row = 4
ws[f"A{row}"] = "Region"
ws[f"B{row}"] = AWS_REGION
row += 1
ws[f"A{row}"] = "How AWS is priced here"
ws[f"B{row}"] = AWS_PRICING_BASIS
row += 2

row = write_section_title(ws, row, "Monthly total at a glance")
glance_headers = [
    "Client Setup",
    "Why This Size",
    "AWS Hosting / Month (USD)",
    "AI — Active Month (USD)",
    "Rough Full System / Month (USD)",
]
for idx, header in enumerate(glance_headers, start=1):
    ws.cell(row=row, column=idx, value=header)
style_header(ws, row, len(glance_headers))
active_ai = next(r["Estimated GPT Cost / Month (USD)"] for r in MONTHLY_AI_BUNDLES if r["Usage Month"] == "Active month")
glance_start = row + 1
for i, scenario in enumerate(CLIENT_SCENARIOS):
    hosting = aws_total(scenario)
    ws.cell(row=glance_start + i, column=1, value=scenario["Client Setup"])
    ws.cell(row=glance_start + i, column=2, value=scenario["Why This Size"])
    ws.cell(row=glance_start + i, column=3, value=hosting)
    ws.cell(row=glance_start + i, column=4, value=active_ai)
    ws.cell(row=glance_start + i, column=5, value=round(hosting + active_ai, 2))
style_table(ws, glance_start, glance_start + len(CLIENT_SCENARIOS) - 1, len(glance_headers))
row = glance_start + len(CLIENT_SCENARIOS) + 2

row = write_section_title(ws, row, "AWS hosting breakdown (monthly USD)")
aws_headers = ["Client Setup", "Cost Item", "What It Covers", "Monthly Cost (USD)"]
for idx, header in enumerate(aws_headers, start=1):
    ws.cell(row=row, column=idx, value=header)
style_header(ws, row, len(aws_headers))
aws_start = row + 1
aws_row = aws_start
total_rows: set[int] = set()
for scenario in CLIENT_SCENARIOS:
    setup = scenario["Client Setup"]
    for line in scenario["aws_lines"]:
        ws.cell(row=aws_row, column=1, value=setup)
        ws.cell(row=aws_row, column=2, value=line["Cost Item"])
        ws.cell(row=aws_row, column=3, value=line["What It Covers"])
        ws.cell(row=aws_row, column=4, value=line["Monthly Cost (USD)"])
        if line["Cost Item"] == "AWS hosting total":
            total_rows.add(aws_row)
        aws_row += 1
style_table(ws, aws_start, aws_row - 1, len(aws_headers), total_rows)
row = aws_row + 2

row = write_section_title(ws, row, "AI model costs (GPT-5 mini)")
ws[f"A{row}"] = "Recommended model"
ws[f"B{row}"] = GPT_MODEL
row += 2
ai_headers = ["AI Use", "What it covers", "Estimated Cost Per Case (USD)"]
for idx, header in enumerate(ai_headers, start=1):
    ws.cell(row=row, column=idx, value=header)
style_header(ws, row, len(ai_headers))
ai_start = row + 1
for i, line in enumerate(AI_LINES):
    ws.cell(row=ai_start + i, column=1, value=line["AI Use"])
    ws.cell(row=ai_start + i, column=2, value=line["What it covers"])
    ws.cell(row=ai_start + i, column=3, value=line["Estimated Cost"])
style_table(ws, ai_start, ai_start + len(AI_LINES) - 1, len(ai_headers))
row = ai_start + len(AI_LINES) + 2

row = write_section_title(ws, row, "Simple monthly AI cost examples")
bundle_headers = ["Usage Month", "Assumed Activity", "Estimated GPT Cost / Month (USD)"]
for idx, header in enumerate(bundle_headers, start=1):
    ws.cell(row=row, column=idx, value=header)
style_header(ws, row, len(bundle_headers))
bundle_start = row + 1
for i, bundle in enumerate(MONTHLY_AI_BUNDLES):
    for c, header in enumerate(bundle_headers, start=1):
        ws.cell(row=bundle_start + i, column=c, value=bundle[header])
style_table(ws, bundle_start, bundle_start + len(MONTHLY_AI_BUNDLES) - 1, len(bundle_headers))

fit_columns(ws, {1: 44, 2: 18, 3: 42, 4: 22, 5: 22})

ws2 = wb.create_sheet("Notes")
ws2["A1"] = "Notes"
ws2["A1"].font = Font(size=16, bold=True)
notes = [
    "AWS hosting is priced as one server running the full system — portal, API, database, and files together.",
    "This matches how Arcus is deployed today (Docker on a single VPS), not a split enterprise AWS setup.",
    "No separate managed database, load balancer, or object storage line items — those are not needed for this deployment.",
    "AWS figures use on-demand list prices for US East (N. Virginia), Linux.",
    "Optional extras not included: automated off-site backups, heavy outbound data transfer, and email sending.",
    "With a 1-year AWS savings plan, hosting can often be 30-40% lower than on-demand.",
    "All AI estimates use one model only: GPT-5 mini.",
    "The 'Rough Full System / Month' total combines AWS hosting with a typical active-month AI usage example.",
]
for idx, note in enumerate(notes, start=3):
    ws2[f"A{idx}"] = f"- {note}"
fit_columns(ws2, {1: 120})
for r in ws2.iter_rows(min_row=1, max_row=20, min_col=1, max_col=1):
    for cell in r:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT)
print(f"Wrote {OUT}")
