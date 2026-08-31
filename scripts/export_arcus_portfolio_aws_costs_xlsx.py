from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "arcus-portfolio-aws-costs.xlsx"

AWS_REGION = "AWS US East (N. Virginia)"
AWS_PRICING_BASIS = "On-demand, Linux — one server running the full system (portal, API, database, files)"

# Practical single-server AWS: m5.large us-east-1 ($70.08/mo) + EBS gp3.
EC2_M5_LARGE_MONTHLY = 70.08
EBS_GP3_GB = 0.08
STORAGE_GB = 120  # investee files, application documents, generated reports

CLIENT_PROFILE = {
    "Deployment": "Portfolio module — full system on AWS",
    "Portfolio applications": "300 per month",
    "Investee companies": "24",
    "LP portal clients": "10 (all with portal access)",
    "Reports": "All reports enabled",
    "Why this server size": (
        "Small LP base (10 users) with moderate application volume (300/month) "
        "and 24 investee companies. One server is enough for portal logins, "
        "application processing, investee data, and report generation."
    ),
}


def aws_hosting_lines(compute_monthly: float, storage_gb: int) -> list[dict[str, str | float]]:
    storage = round(storage_gb * EBS_GP3_GB, 2)
    lines = [
        {
            "Cost Item": "Server",
            "What It Covers": "Runs the full Arcus system — portfolio module, LP portal, API, and database",
            "Monthly Cost (USD)": compute_monthly,
        },
        {
            "Cost Item": "Storage",
            "What It Covers": "Application files, investee documents, database, and generated reports",
            "Monthly Cost (USD)": storage,
        },
    ]
    total = round(compute_monthly + storage, 2)
    lines.append(
        {
            "Cost Item": "AWS hosting total",
            "What It Covers": "Estimated monthly cost to run the full system on AWS",
            "Monthly Cost (USD)": total,
        }
    )
    return lines


AWS_LINES = aws_hosting_lines(EC2_M5_LARGE_MONTHLY, STORAGE_GB)
AWS_TOTAL = next(row["Monthly Cost (USD)"] for row in AWS_LINES if row["Cost Item"] == "AWS hosting total")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="E8F0FE")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row_num: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_table(ws, start_row: int, end_row: int, cols: int, total_rows: set[int] | None = None):
    total_rows = total_rows or set()
    for row_idx in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if row_idx in total_rows:
                cell.fill = TOTAL_FILL
                cell.font = Font(bold=True)


def fit_columns(ws, widths: dict[int, int]):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_section_title(ws, row: int, title: str) -> int:
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = Font(size=13, bold=True)
    return row + 1


wb = Workbook()

ws = wb.active
ws.title = "AWS Hosting Costs"
ws["A1"] = "Arcus Portfolio — AWS Server Cost Estimate"
ws["A1"].font = Font(size=16, bold=True)
ws["A2"] = "Monthly AWS hosting only. AI costs are not included."
ws["A2"].alignment = Alignment(wrap_text=True)

row = 4
ws[f"A{row}"] = "Region"
ws[f"B{row}"] = AWS_REGION
row += 1
ws[f"A{row}"] = "How AWS is priced here"
ws[f"B{row}"] = AWS_PRICING_BASIS
row += 2

row = write_section_title(ws, row, "Client setup")
profile_headers = ["Item", "Value"]
for idx, header in enumerate(profile_headers, start=1):
    ws.cell(row=row, column=idx, value=header)
style_header(ws, row, len(profile_headers))
profile_start = row + 1
profile_rows = [
    ("Deployment", CLIENT_PROFILE["Deployment"]),
    ("Portfolio applications", CLIENT_PROFILE["Portfolio applications"]),
    ("Investee companies", CLIENT_PROFILE["Investee companies"]),
    ("LP portal clients", CLIENT_PROFILE["LP portal clients"]),
    ("Reports", CLIENT_PROFILE["Reports"]),
    ("Why this server size", CLIENT_PROFILE["Why this server size"]),
]
for i, (label, value) in enumerate(profile_rows):
    ws.cell(row=profile_start + i, column=1, value=label)
    ws.cell(row=profile_start + i, column=2, value=value)
style_table(ws, profile_start, profile_start + len(profile_rows) - 1, len(profile_headers))
row = profile_start + len(profile_rows) + 2

row = write_section_title(ws, row, "Monthly AWS hosting total")
ws[f"A{row}"] = "Estimated AWS hosting / month (USD)"
ws[f"B{row}"] = AWS_TOTAL
ws[f"B{row}"].font = Font(size=14, bold=True)
row += 2

row = write_section_title(ws, row, "AWS hosting breakdown (monthly USD)")
aws_headers = ["Cost Item", "What It Covers", "Monthly Cost (USD)"]
for idx, header in enumerate(aws_headers, start=1):
    ws.cell(row=row, column=idx, value=header)
style_header(ws, row, len(aws_headers))
aws_start = row + 1
total_rows: set[int] = set()
for i, line in enumerate(AWS_LINES):
    r = aws_start + i
    ws.cell(row=r, column=1, value=line["Cost Item"])
    ws.cell(row=r, column=2, value=line["What It Covers"])
    ws.cell(row=r, column=3, value=line["Monthly Cost (USD)"])
    if line["Cost Item"] == "AWS hosting total":
        total_rows.add(r)
style_table(ws, aws_start, aws_start + len(AWS_LINES) - 1, len(aws_headers), total_rows)

fit_columns(ws, {1: 28, 2: 58, 3: 22})

ws2 = wb.create_sheet("Notes")
ws2["A1"] = "Notes"
ws2["A1"].font = Font(size=16, bold=True)
notes = [
    "This workbook covers AWS server hosting only. AI / GPT costs are excluded.",
    "Client profile: portfolio module, 300 applications per month, 24 investee companies, 10 LP portal users, all reports.",
    "AWS hosting is one server running the full system — portal, API, database, and files together.",
    "This matches a standard single-server deployment (same approach as a VPS).",
    "AWS figures use on-demand list prices for US East (N. Virginia), Linux.",
    "Optional extras not included: automated off-site backups, heavy outbound data transfer, and email sending.",
    "With a 1-year AWS savings plan, hosting can often be 30-40% lower than on-demand.",
]
for idx, note in enumerate(notes, start=3):
    ws2[f"A{idx}"] = f"- {note}"
fit_columns(ws2, {1: 120})
for r in ws2.iter_rows(min_row=1, max_row=15, min_col=1, max_col=1):
    for cell in r:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT)
print(f"Wrote {OUT}")
